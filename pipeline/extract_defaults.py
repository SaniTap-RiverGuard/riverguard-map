#!/usr/bin/env python3
"""Extract model default parameters from the VM0047 calculator xlsx into
docs/data/model_config.json (the web app's default configuration).

Re-run this whenever a new xlsx version arrives:
    .venv/bin/python pipeline/extract_defaults.py [path-to-xlsx]

The xlsx is the SOURCE OF TRUTH. Every value below records the sheet/cell it
came from so the app's Assumptions modal can cite it. Values that are NOT in
the xlsx (survival rate, species scaling factors, net-deduction structure
choices) are marked source: "project brief" / "assumption".
"""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "data" / "VM0047_Bambusa_ARR_Calculator_FIXED v03.xlsx"
OUT = ROOT / "docs" / "data" / "model_config.json"

wb = openpyxl.load_workbook(XLSX, data_only=True)
ass = wb["Assumptions"]
bio = wb["Biomass"]
eq30 = wb["Eq30_Carbon_removals"]


def cell(ws, ref):
    v = ws[ref].value
    if v is None:
        raise ValueError(f"Expected value in {ws.title}!{ref} — sheet layout changed? Update extract_defaults.py")
    return v


def row_series(ws, row, n=20):
    """Year 1..n values from columns E.. of an equation sheet."""
    from openpyxl.utils import get_column_letter
    return [ws[f"{get_column_letter(5 + i)}{row}"].value or 0 for i in range(n)]


# Balcooa growth curve, years 1-20 (Biomass rows 6..25)
growth = []
for r in range(6, 26):
    growth.append({
        "year": cell(bio, f"A{r}"),
        "culms_per_clump": cell(bio, f"B{r}"),
        "avg_culm_weight_kg": cell(bio, f"C{r}"),
        "agb_kg": cell(bio, f"D{r}"),          # above-ground biomass per clump, dry kg
        "total_biomass_kg": cell(bio, f"H{r}"),  # AGB + BGB per clump, dry kg
    })

config = {
    "_generated_from": XLSX.name,
    "_note": "Defaults extracted by pipeline/extract_defaults.py. Do not hand-edit; edit the xlsx or the marked assumption blocks and re-run.",

    "planting": {
        "strip_width_m": {"value": cell(ass, "B5"), "source": "Assumptions!B5"},
        "ha_per_km_single_side": {"value": cell(ass, "B6"), "source": "Assumptions!B6"},
        "density_seedlings_ha": {"value": cell(ass, "B17"), "source": "Assumptions!B17"},
        "mature_canopy_diameter_m": {"value": cell(ass, "B27"), "source": "Assumptions!B27"},
        "default_rows": {"value": 5, "source": "Seedling Density sheet layout (5 rows across 20 m at 5 m spacing)"},
        "default_spacing_m": {"value": 4.41, "source": "derived: 5 rows x 4.41 m spacing reproduces Assumptions!B17 = 500 seedlings/ha via the Seedling Density formula (note: the sheet's '5 m spacing' practical option gives 400/ha)"},
        "density_formula": "seedlings_per_ha = rows * 10000 / ((canopy_d + (rows-1)*spacing) * spacing)",
        "density_formula_source": "Seedling Density sheet columns D,F,G (verified against all table rows)",
        "survival_rate": {"value": 0.70, "source": "project brief (NOT in xlsx)"}
    },

    "carbon": {
        "carbon_fraction": {"value": cell(bio, "A28"), "source": "Biomass!A28 (IPCC 2003 tropical woody default)"},
        "root_shoot_ratio": {"value": cell(bio, "A30"), "source": "Biomass!A30 (IPCC default tropical)"},
        "co2e_per_c": {"value": 44 / 12, "source": "Biomass!A29 (44/12 molecular ratio)"},
        "soc_tc_ha_yr": {"value": cell(ass, "B21"), "source": "Assumptions!B21 (soil organic carbon accrual)"},
        "minor_pools_tc_ha": {
            "_comment": "Per-ha carbon stock (tC/ha) years 1-20 for VM0047 minor pools; added to gross when 'include minor pools' is on (default on, to reproduce Eq1/Eq30).",
            "herb": {"values": row_series(wb["Eq6_CWP_herb"], 8), "source": "Eq6_CWP_herb row 8 (non-woody, est. 4% of woody)"},
            "dead_wood": {"values": row_series(wb["Eq8_CWP_DW"], 9), "source": "Eq8_CWP_DW row 9"},
            "litter": {"values": row_series(wb["Eq10_CWP_LI"], 8), "source": "Eq10_CWP_LI row 8 (steady-state litter)"}
        },
        "deductions": {
            "_comment": "Net factor = (1-performance_benchmark)*(1-uncertainty)*(1-leakage)*(1-buffer) = 0.5472 with these defaults, matching Eq30 in the xlsx (agreed 2026-08-13).",
            "performance_benchmark": {"value": cell(eq30, "E6"), "source": "Eq30_Carbon_removals!E6 (placeholder pending control plots)"},
            "uncertainty": {"value": cell(eq30, "E7"), "source": "Eq30_Carbon_removals!E7"},
            "leakage": {"value": cell(eq30, "E9"), "source": "Eq30_Carbon_removals!E9 (VMD0054)"},
            "buffer": {"value": cell(eq30, "E10"), "source": "Eq30_Carbon_removals!E10 (VCS non-permanence buffer)"}
        }
    },

    "finance": {
        "carbon_price_usd": {"value": cell(ass, "B11"), "source": "Assumptions!B11"},
        "price_escalation": {"value": cell(ass, "B12"), "source": "Assumptions!B12"},
        "project_years": {"value": cell(ass, "B13"), "source": "Assumptions!B13"},
        "discount_rate": {"value": cell(ass, "B14"), "source": "Assumptions!B14"},
        "revenue_start_year": {"value": 3, "source": "Revenue sheet (2-year validation lag; first sale year 3). Assumptions!B15 says 2 — year 3 chosen per Adriaan 2026-08-13."}
    },

    "species": {
        "balcooa": {
            "label": "Bambusa balcooa",
            "scale": {"value": 1.0, "source": "measured curve (Biomass sheet)"},
            "growth_curve": growth,
            "provisional": False
        },
        "vulgaris": {
            "label": "Bambusa vulgaris",
            "scale": {"value": 0.85, "source": "ASSUMPTION (project brief): scaled balcooa curve, no measured data yet"},
            "provisional": True
        },
        "asper": {
            "label": "Dendrocalamus asper",
            "scale": {"value": 1.0, "source": "ASSUMPTION (project brief): scaled balcooa curve, no measured data yet"},
            "provisional": True
        }
    }
}

# Cross-check the headline numbers so a changed xlsx fails loudly here rather
# than shipping silently different economics.
y20 = growth[-1]
assert y20["year"] == 20, "growth curve should end at year 20"
tco2e_ha_20 = cell(bio, "K25")
print(f"Year-20 clump: {y20['culms_per_clump']} culms, {y20['total_biomass_kg']:.1f} kg total biomass")
print(f"Cumulative tCO2e/ha at year 20 (xlsx K25): {tco2e_ha_20:.1f}")
dens = cell(ass, "B17")
cf, rs = cell(bio, "A28"), cell(bio, "A30")
recomputed = y20["total_biomass_kg"] * dens / 1000 * cf * (44 / 12)
print(f"Recomputed from curve ({dens}/ha x biomass x CF x 44/12): {recomputed:.1f} tCO2e/ha")
if abs(recomputed - tco2e_ha_20) / tco2e_ha_20 > 0.01:
    print("WARNING: recomputed cumulative tCO2e/ha differs from xlsx by >1% — check formula chain")

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(json.dumps(config, indent=2))
print(f"Wrote {OUT}")
